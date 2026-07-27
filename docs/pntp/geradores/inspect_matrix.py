from pathlib import Path

from openpyxl import load_workbook


matrix_path = Path(__file__).resolve().parents[1] / "docs" / "pntp" / "matriz-criterios-2025.xlsx"
workbook = load_workbook(matrix_path, data_only=True)
worksheet = workbook.active
for row in worksheet.iter_rows(min_row=2, values_only=True):
    matrix, dimension, criterion_id, criterion, classification = row[:5]
    if not matrix:
        continue
    normalized = str(matrix).strip().upper()
    if normalized in {"PODER LEGISLATIVO", "COMUM"}:
        if normalized == "PODER LEGISLATIVO" or dimension in {
            "Receita",
            "Despesa",
            "Recursos Humanos",
            "Licitações",
            "Contratos",
            "Diárias",
            "Serviço de Informações ao Cidadão (SIC)",
            "Acessibilidade",
            "Ouvidoria",
            "LGPD e Governo Digital",
        }:
            print(
                f"{normalized} | {dimension} | {criterion_id} | "
                f"{classification} | {criterion}"
            )
