# -*- coding: utf-8 -*-
"""Coleta emendas estaduais destinadas a Varginha na fonte oficial da SEGOV-MG.

Fonte:
  https://www.emendas.mg.gov.br/transparencia/

O Portal de Emendas publica uma planilha consolidada extraída de SIGCON-MG,
SIAFI-MG e SIAD-MG. Este coletor descobre o XLSX mais recente, preserva uma
cópia privada para auditoria e gera a camada consumida por /emendas/.

Regra central: "Valor Utilizado" NÃO é tratado como empenhado, pago, recebido
ou executado. Cada estágio financeiro é publicado em seu próprio campo.
"""
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import html
import json
import os
import re
import shutil
import sys
import tempfile
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
DESTINO = BASE / "data" / "emendas_estaduais_normalizadas.js"
ROOT = BASE.parents[1]
CACHE_DIR = ROOT / "private" / "cache" / "emendas-estaduais"
CACHE_XLSX = CACHE_DIR / "dados-oficiais-latest.xlsx"
CACHE_META = CACHE_DIR / "metadados-latest.json"
HISTORY_META = CACHE_DIR / "historico-coletas.json"

FONTE_PAGINA = "https://www.emendas.mg.gov.br/transparencia/"
IBGE_VARGINHA = "3170701"
USER_AGENT = "FiscalizaVarginha/1.0 (+https://fiscalizavarginha.com.br)"
RETRYABLE_HTTP = {429, 500, 502, 503, 504}

COLUNAS_OBRIGATORIAS = {
    "ano da indicacao",
    "numero da indicacao",
    "status da indicacao",
    "autor",
    "codigo ibge do municipio",
    "municipio",
    "nome beneficiario",
    "valor indicado",
    "valor utilizado",
    "valor empenhado no ano",
    "valor liquidado atualizado",
    "valor pago atualizado",
    "valor executado",
    "status",
}


def agora_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).astimezone().isoformat(timespec="seconds")


def normalizar_chave(value: Any) -> str:
    texto = unicodedata.normalize("NFKD", str(value or ""))
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).strip().lower()


def texto(value: Any) -> str | None:
    if value is None:
        return None
    value = html.unescape(str(value)).replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value).strip()
    return None if not value or value == "-" else value


def numero_inteiro_texto(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)) and float(value).is_integer():
        return str(int(value))
    value = texto(value)
    return re.sub(r"\.0$", "", value) if value else None


def dinheiro(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        number = float(value)
    else:
        raw = str(value).strip().replace("R$", "").replace(" ", "")
        if not raw or raw == "-":
            return None
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        try:
            number = float(raw)
        except ValueError:
            return None
    return round(number, 2)


def dinheiro_texto(value: float | None) -> str | None:
    if value is None:
        return None
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def data_texto(value: Any) -> str | None:
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%d/%m/%Y")
    value = texto(value)
    if not value:
        return None
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(value[:10], formato).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return value


def cnpj_texto(value: Any) -> str | None:
    value = numero_inteiro_texto(value)
    if not value:
        return None
    digits = re.sub(r"\D", "", value)
    if len(digits) != 14:
        return value
    return f"{digits[:2]}.{digits[2:5]}.{digits[5:8]}/{digits[8:12]}-{digits[12:]}"


def obter(url: str, timeout: int = 180, attempts: int = 4) -> tuple[bytes, dict[str, str]]:
    last_error: Exception | None = None
    for attempt in range(attempts):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
            with urllib.request.urlopen(req, timeout=timeout) as response:
                headers = {
                    "content_type": response.headers.get("Content-Type", ""),
                    "etag": response.headers.get("ETag", ""),
                    "last_modified": response.headers.get("Last-Modified", ""),
                }
                return response.read(), headers
        except urllib.error.HTTPError as exc:
            last_error = exc
            if exc.code not in RETRYABLE_HTTP or attempt == attempts - 1:
                raise
        except (urllib.error.URLError, TimeoutError, OSError) as exc:
            last_error = exc
            if attempt == attempts - 1:
                raise
        delay = min(20, 2**attempt * 1.5)
        print(f"Fonte estadual indisponível; nova tentativa em {delay:.1f}s.")
        time.sleep(delay)
    raise last_error or RuntimeError("Falha de rede sem detalhe")


def descobrir_planilha(page_html: str) -> str:
    hrefs = re.findall(r"""href=["']([^"']+\.xlsx(?:\?[^"']*)?)["']""", page_html, re.I)
    candidatos: list[tuple[int, int, str]] = []
    for href in hrefs:
        url = urllib.parse.urljoin(FONTE_PAGINA, html.unescape(href))
        nome = urllib.parse.unquote(urllib.parse.urlparse(url).path.rsplit("/", 1)[-1])
        anos = [int(year) for year in re.findall(r"20\d{2}", nome)]
        if "DADOS_EMENDAS" not in nome.upper() or not anos:
            continue
        candidatos.append((max(anos), min(anos), url))
    if not candidatos:
        raise ValueError("Planilha consolidada de emendas não encontrada na página oficial")
    return max(candidatos)[2]


def cabecalhos(worksheet) -> tuple[list[str], dict[str, int]]:
    values = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    normalized = [normalizar_chave(value) for value in values]
    index = {name: position for position, name in enumerate(normalized)}
    missing = sorted(COLUNAS_OBRIGATORIAS - set(index))
    if missing:
        raise ValueError("Colunas oficiais ausentes: " + ", ".join(missing))
    return [str(value or "") for value in values], index


def valor_linha(values: tuple[Any, ...], index: dict[str, int], column: str) -> Any:
    position = index.get(normalizar_chave(column))
    return values[position] if position is not None and position < len(values) else None


def estagio_atual(status_indicacao: str | None, stages: dict[str, float | None]) -> str:
    status_key = normalizar_chave(status_indicacao)
    if "reprov" in status_key:
        return "reprovada"
    for key in ("executado", "pago", "liquidado", "empenhado", "utilizado"):
        if (stages.get(key) or 0) > 0:
            return key
    if "aprov" in status_key or "enviada" in status_key:
        return "aprovada"
    return "indicada"


def objeto_da_linha(values: tuple[Any, ...], index: dict[str, int]) -> str | None:
    candidates = [
        valor_linha(values, index, "Descrição da Indicação"),
        valor_linha(values, index, "Especificação"),
        valor_linha(values, index, "Tipo de Aplicação"),
        valor_linha(values, index, "Nome da Ação"),
    ]
    return next((item for item in map(texto, candidates) if item), None)


def normalizar_linha(
    values: tuple[Any, ...],
    index: dict[str, int],
    source_url: str,
    source_hash: str,
    collected_at: str,
    sheet_name: str,
    excel_row: int,
) -> dict[str, Any] | None:
    ibge = numero_inteiro_texto(valor_linha(values, index, "Código IBGE do Município"))
    if ibge != IBGE_VARGINHA:
        return None

    year = numero_inteiro_texto(valor_linha(values, index, "Ano da Indicação"))
    indication = numero_inteiro_texto(valor_linha(values, index, "Número da Indicação"))
    if not year or not indication:
        raise ValueError(f"Linha {excel_row}: ano/número da indicação ausente")

    status_indicacao = texto(valor_linha(values, index, "Status da Indicação"))
    status_financeiro = texto(valor_linha(values, index, "Status"))
    valor_indicado = dinheiro(valor_linha(values, index, "Valor Indicado")) or 0.0
    stages = {
        "utilizado": dinheiro(valor_linha(values, index, "Valor Utilizado")),
        "empenhado": dinheiro(valor_linha(values, index, "Valor Empenhado no Ano")),
        "liquidadoAno": dinheiro(valor_linha(values, index, "Valor Liquidado no Ano")),
        "liquidado": dinheiro(valor_linha(values, index, "Valor Liquidado Atualizado")),
        "pagoAno": dinheiro(valor_linha(values, index, "Valor Pago no Ano")),
        "pago": dinheiro(valor_linha(values, index, "Valor Pago Atualizado")),
        "executado": dinheiro(valor_linha(values, index, "Valor Executado")),
        "restosInscritos": dinheiro(valor_linha(values, index, "Valor Inscrito em Restos a Pagar")),
        "restosSaldo": dinheiro(valor_linha(values, index, "Saldo Restos a Pagar")),
    }
    rejected = "reprov" in normalizar_chave(status_indicacao or status_financeiro)
    tipo_indicacao = texto(valor_linha(values, index, "Tipo de Indicação"))
    categoria = texto(valor_linha(values, index, "Categoria"))
    impositiva = normalizar_chave(valor_linha(values, index, "Indicador de Impositividade")) == "s"
    beneficiary = texto(valor_linha(values, index, "Nome Beneficiário"))
    source_record_url = f"{source_url}#planilha={urllib.parse.quote(sheet_name)}&linha={excel_row}"

    return {
        "id": f"estadual-sigcon-{year}-{indication}",
        "tipo": "Estadual",
        "esferaDocumento": "Estadual",
        "granularidade": "indicacao_estadual_oficial",
        "ano": year,
        "anoEmenda": year,
        "anoRecurso": year if (stages["pago"] or 0) > 0 else None,
        "emenda": f"{indication}/{year}",
        "emendaOriginal": indication,
        "numeroIndicacao": indication,
        "origemExternaId": f"sigcon-{indication}",
        "autor": (texto(valor_linha(values, index, "Autor")) or "").upper(),
        "cargoAutor": "Dep. Estadual / autoria coletiva",
        "tipoIndicacao": tipo_indicacao,
        "categoria": categoria or tipo_indicacao,
        "modalidade": tipo_indicacao,
        "impositiva": impositiva,
        "beneficiario": beneficiary or "",
        "documentoBeneficiario": cnpj_texto(valor_linha(values, index, "Número do CNPJ do Beneficiário")),
        "codigoIbge": IBGE_VARGINHA,
        "localidade": texto(valor_linha(values, index, "Município")) or "VARGINHA",
        "orgao": texto(valor_linha(values, index, "Unidade Orçamentária Descrição")),
        "orgaoSigla": texto(valor_linha(values, index, "Unidade Orçamentária Sigla")),
        "funcao": texto(valor_linha(values, index, "Função Descrição")),
        "acao": texto(valor_linha(values, index, "Nome da Ação")),
        "grupoDespesa": texto(valor_linha(values, index, "Grupo de Despesa Descrição")),
        "objeto": objeto_da_linha(values, index),
        "descricao": texto(valor_linha(values, index, "Descrição da Indicação")),
        "valor": valor_indicado,
        "valorTexto": dinheiro_texto(valor_indicado),
        "valorDeclarado": valor_indicado,
        "valorIndicado": valor_indicado,
        "valorUtilizado": stages["utilizado"],
        "valorEmpenhado": stages["empenhado"],
        "valorLiquidadoAno": stages["liquidadoAno"],
        "valorLiquidado": stages["liquidado"],
        "valorPagoAno": stages["pagoAno"],
        "valorPago": stages["pago"],
        "valorExecutado": stages["executado"],
        "valorRestosInscritos": stages["restosInscritos"],
        "valorRestosSaldo": stages["restosSaldo"],
        "valorRecebido": None,
        "identificador_repasse_confirmado": False,
        "statusIndicacao": status_indicacao,
        "statusFinanceiro": status_financeiro,
        "estagioAtual": estagio_atual(status_indicacao, stages),
        "aprovado": "Não" if rejected else "Sim",
        "justificativaReprovacao": texto(valor_linha(values, index, "Justificativa de Reprovação")),
        "classificacaoIot": texto(valor_linha(values, index, "Classificação do IOT")),
        "numeroInstrumento": numero_inteiro_texto(valor_linha(values, index, "Número do Instrumento")),
        "statusInstrumento": texto(valor_linha(values, index, "Status do Instrumento")),
        "codigoSiafiInstrumento": numero_inteiro_texto(valor_linha(values, index, "Código SIAFI do Instrumento")),
        "dataPublicacaoInstrumento": data_texto(valor_linha(values, index, "Data de Publicação do Instrumento")),
        "dataValidadeInstrumento": data_texto(valor_linha(values, index, "Data de Validade do Instrumento")),
        "classificacaoComprovacao": "confirmado",
        "fonte": "Portal de Emendas de Minas Gerais — SIGCON-MG / SIAFI-MG / SIAD-MG",
        "fonteUrl": source_url,
        "fonteRegistroUrl": source_record_url,
        "evidencias": {
            "urlOficial": source_url,
            "paginaOficial": FONTE_PAGINA,
            "arquivoSha256": source_hash,
            "planilha": sheet_name,
            "linha": excel_row,
            "coletadoEm": collected_at,
            "estagiosSemInferencia": True,
        },
        "fontes": [{
            "nome": "Relatório de Execução Geral — Portal de Emendas MG",
            "url": source_url,
            "sha256": source_hash,
            "planilha": sheet_name,
            "linha": excel_row,
        }],
        "pendencias": ([] if beneficiary else ["beneficiario_nao_informado"])
        + ([] if objeto_da_linha(values, index) else ["objeto_nao_detalhado"]),
    }


def ler_planilha(
    path: Path,
    source_url: str,
    source_hash: str,
    collected_at: str,
) -> tuple[list[dict[str, Any]], str, list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if not workbook.sheetnames:
        raise ValueError("XLSX oficial sem planilhas")
    sheet_name = workbook.sheetnames[0]
    worksheet = workbook[sheet_name]
    headers, index = cabecalhos(worksheet)
    records: list[dict[str, Any]] = []
    for excel_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        record = normalizar_linha(
            values, index, source_url, source_hash, collected_at, sheet_name, excel_row,
        )
        if record:
            records.append(record)
    workbook.close()
    if not records:
        raise ValueError("Nenhuma emenda de Varginha encontrada no XLSX oficial")
    ids = [record["id"] for record in records]
    if len(ids) != len(set(ids)):
        duplicates = [key for key, count in Counter(ids).items() if count > 1]
        raise ValueError(f"IDs estaduais duplicados: {duplicates[:5]}")
    records.sort(key=lambda item: (-int(item["ano"]), -int(item["numeroIndicacao"])))
    return records, sheet_name, headers


def somar(records: Iterable[dict[str, Any]], field: str) -> float:
    return round(sum(float(record.get(field) or 0) for record in records), 2)


def construir_payload(
    records: list[dict[str, Any]],
    source_url: str,
    source_hash: str,
    collected_at: str,
    sheet_name: str,
    headers: list[str],
    http_headers: dict[str, str],
) -> dict[str, Any]:
    by_year: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for record in records:
        by_year[record["ano"]].append(record)
    year_summary = {}
    for year, items in sorted(by_year.items()):
        year_summary[year] = {
            "registros": len(items),
            "valorIndicado": somar(items, "valorIndicado"),
            "valorUtilizado": somar(items, "valorUtilizado"),
            "valorEmpenhado": somar(items, "valorEmpenhado"),
            "valorLiquidado": somar(items, "valorLiquidado"),
            "valorPago": somar(items, "valorPago"),
            "valorExecutado": somar(items, "valorExecutado"),
        }
    statuses = Counter(record["statusIndicacao"] or "Não informado" for record in records)
    return {
        "metadata": {
            "criterio": "Registros oficiais por indicação; cada estágio financeiro é mantido sem inferência.",
            "fonte": "Portal de Emendas de Minas Gerais — SIGCON-MG / SIAFI-MG / SIAD-MG",
            "fontePagina": FONTE_PAGINA,
            "fonteArquivo": source_url,
            "arquivoSha256": source_hash,
            "arquivoEtag": http_headers.get("etag") or None,
            "arquivoUltimaModificacao": http_headers.get("last_modified") or None,
            "extraidoEm": collected_at,
            "codigoIbge": IBGE_VARGINHA,
            "planilha": sheet_name,
            "colunasOriginais": headers,
            "totalRegistros": len(records),
            "anos": year_summary,
            "statusIndicacao": dict(sorted(statuses.items())),
            "totalIndicado": somar(records, "valorIndicado"),
            "totalUtilizado": somar(records, "valorUtilizado"),
            "totalEmpenhado": somar(records, "valorEmpenhado"),
            "totalLiquidado": somar(records, "valorLiquidado"),
            "totalPago": somar(records, "valorPago"),
            "totalExecutado": somar(records, "valorExecutado"),
            "registrosComPagamento": sum((record.get("valorPago") or 0) > 0 for record in records),
            "registrosSemPagamento": sum((record.get("valorPago") or 0) == 0 for record in records),
            "observacao": (
                "Valor utilizado representa recurso reservado/associado à indicação e não comprova pagamento. "
                "Pagamento e execução usam exclusivamente as colunas oficiais correspondentes."
            ),
        },
        "emendas": records,
    }


def serializar(payload: dict[str, Any]) -> str:
    return (
        "window.EMENDAS_ESTADUAIS_NORMALIZADAS = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n"
    )


def escrever_atomico(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=path.name + ".", suffix=".tmp", dir=path.parent)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temp_name, path)
    finally:
        if os.path.exists(temp_name):
            os.unlink(temp_name)


def atualizar_historico(metadata: dict[str, Any]) -> None:
    history: list[dict[str, Any]] = []
    if HISTORY_META.exists():
        try:
            history = json.loads(HISTORY_META.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            history = []
    event = {
        "extraidoEm": metadata["extraidoEm"],
        "arquivoSha256": metadata["arquivoSha256"],
        "fonteArquivo": metadata["fonteArquivo"],
        "totalRegistros": metadata["totalRegistros"],
        "totalIndicado": metadata["totalIndicado"],
        "totalPago": metadata["totalPago"],
        "anos": metadata["anos"],
    }
    if not history or history[-1].get("arquivoSha256") != event["arquivoSha256"]:
        history.append(event)
    history = history[-90:]
    escrever_atomico(HISTORY_META, json.dumps(history, ensure_ascii=False, indent=2) + "\n")


def coletar(input_path: Path | None = None) -> dict[str, Any]:
    collected_at = agora_iso()
    http_headers: dict[str, str] = {}
    source_url = ""
    temp_path: Path | None = None
    try:
        if input_path:
            source_url = input_path.resolve().as_uri()
            xlsx_path = input_path
            source_bytes = input_path.read_bytes()
        else:
            page_bytes, _ = obter(FONTE_PAGINA, timeout=60)
            page_html = page_bytes.decode("utf-8", errors="replace")
            source_url = descobrir_planilha(page_html)
            source_bytes, http_headers = obter(source_url)
            fd, temp_name = tempfile.mkstemp(prefix="emendas-mg-", suffix=".xlsx")
            os.close(fd)
            temp_path = Path(temp_name)
            temp_path.write_bytes(source_bytes)
            xlsx_path = temp_path

        source_hash = hashlib.sha256(source_bytes).hexdigest()
        records, sheet_name, headers = ler_planilha(
            xlsx_path, source_url, source_hash, collected_at,
        )
        payload = construir_payload(
            records, source_url, source_hash, collected_at, sheet_name, headers, http_headers,
        )

        escrever_atomico(DESTINO, serializar(payload))
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        if input_path:
            shutil.copyfile(input_path, CACHE_XLSX)
        else:
            CACHE_XLSX.write_bytes(source_bytes)
        escrever_atomico(CACHE_META, json.dumps(payload["metadata"], ensure_ascii=False, indent=2) + "\n")
        atualizar_historico(payload["metadata"])
        return payload
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, help="XLSX local para reprodução/testes")
    parser.add_argument("--strict", action="store_true", help="Falhar mesmo se existir base anterior")
    args = parser.parse_args()
    try:
        payload = coletar(args.input)
    except Exception as exc:
        if DESTINO.exists() and not args.strict:
            print(
                f"AVISO: coleta estadual falhou ({exc}); base anterior preservada.",
                file=sys.stderr,
            )
            return 2
        raise
    print(json.dumps(payload["metadata"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
